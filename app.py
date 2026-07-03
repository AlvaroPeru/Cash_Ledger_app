import streamlit as st
import pytz
import openpyxl
import io
import os
from datetime import datetime
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table,
    TableStyle, PageBreak, HRFlowable
)

MONTHS = {
    1:"January", 2:"February", 3:"March", 4:"April",
    5:"May",     6:"June",     7:"July",  8:"August",
    9:"September",10:"October",11:"November",12:"December"
}

INCOME_TYPES  = {"cash income", "funding"}
EXPENSE_TYPES = {"expense", "return cash"}

DARK  = colors.HexColor("#1a1a2e")
GBGD  = colors.HexColor("#f5f5f5")
GBRD  = colors.HexColor("#e8e8e8")
MID   = colors.HexColor("#666670")
RED   = colors.HexColor("#a32d2d")
GREEN = colors.HexColor("#0f6e56")


# ─── HELPERS ────────────────────────────────

def _fmt(v):
    try:    return f"${float(v):,.2f}"
    except: return ""

def _fecha(r):
    try:    return f"{int(r['month']):02d}/{int(r['day']):02d}/{r['year']}"
    except: return ""

def _prop(p):
    return p.replace("Property 1: ","").replace("Property 2: ","")

def fmt_local(r):
    try:
        v = float(r["local"] or 0)
        cur = r["cur"].strip()
        return f"${v:,.2f}" if cur == "USD" else f"{v:,.0f} {cur}"
    except:
        return ""

def is_income(r):
    return str(r.get("type","")).strip().lower() in INCOME_TYPES

def is_expense(r):
    return str(r.get("type","")).strip().lower() in EXPENSE_TYPES


# ─── LECTURA ────────────────────────────────

def leer_excel(file):
    wb = openpyxl.load_workbook(file, data_only=True)
    ws = next((wb[s] for s in wb.sheetnames if "cash" in s.lower()), wb.active)
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("The file appears to be empty.")

    header_idx = 0
    for i, r in enumerate(rows):
        if any(str(c or "").strip().lower() in ("year","month","date","description") for c in r):
            header_idx = i
            break
    rows = rows[header_idx:]
    hdrs = [str(h or "").strip().lower() for h in rows[0]]

    def col(name):
        for i, h in enumerate(hdrs):
            if name in h:
                return i
        return None

    ci = dict(
        year=col("year"), month=col("month"), day=col("day"),
        desc=col("description"), local=col("local amount"),
        usd=col("usd amount"), bal=col("running balance"),
        cat=col("category"), cur=col("currency"),
        prop=col("property"), type=col("type"),
    )

    data = []
    for r in rows[1:]:
        if ci["year"] is None or ci["month"] is None:
            break
        if not r[ci["year"]] or not r[ci["month"]]:
            continue
        try:
            mes  = int(r[ci["month"]])
            anio = int(r[ci["year"]])
        except:
            continue
        data.append({
            "year":  anio, "month": mes,
            "day":   r[ci["day"]]   if ci["day"]   is not None else "",
            "desc":  str(r[ci["desc"]]  or "").strip() if ci["desc"]  is not None else "",
            "local": r[ci["local"]] if ci["local"] is not None else None,
            "usd":   r[ci["usd"]]   if ci["usd"]   is not None else None,
            "bal":   r[ci["bal"]]   if ci["bal"]   is not None else None,
            "cat":   str(r[ci["cat"]]   or "").strip() if ci["cat"]   is not None else "",
            "cur":   str(r[ci["cur"]]   or "").strip() if ci["cur"]   is not None else "",
            "prop":  str(r[ci["prop"]]  or "").strip() if ci["prop"]  is not None else "",
            "type":  str(r[ci["type"]]  or "").strip() if ci["type"]  is not None else "",
        })
    return data


def agrupar(data):
    g = {}
    for r in data:
        k = (r["year"], r["month"])
        g.setdefault(k, []).append(r)
    return dict(sorted(g.items()))


# ─── TABLE BUILDER ──────────────────────────

def make_tx_table(rows, header_color=None):
    hc = header_color or DARK
    heads = ["Date", "Prop.", "Category", "Description", "Cur.", "Local Amount", "USD"]
    col_w = [0.7*inch, 0.55*inch, 1.0*inch, 2.9*inch, 0.38*inch, 0.8*inch, 0.67*inch]

    data = [heads] + [[
        _fecha(r),
        _prop(r["prop"])[:6],
        r["cat"][:16],
        r["desc"][:60],
        r["cur"],
        fmt_local(r),
        f'${float(r["usd"] or 0):,.2f}',
    ] for r in rows]

    t = Table(data, colWidths=col_w, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,0),  hc),
        ("TEXTCOLOR",     (0,0),(-1,0),  colors.white),
        ("FONTNAME",      (0,0),(-1,0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0,0),(-1,0),  7.5),
        ("ALIGN",         (0,0),(-1,0),  "CENTER"),
        ("FONTNAME",      (0,1),(-1,-1), "Helvetica"),
        ("FONTSIZE",      (0,1),(-1,-1), 7.5),
        ("ROWBACKGROUNDS",(0,1),(-1,-1), [colors.white, GBGD]),
        ("TEXTCOLOR",     (5,1),(5,-1),  MID),
        ("ALIGN",         (5,1),(5,-1),  "RIGHT"),
        ("TEXTCOLOR",     (6,1),(6,-1),  RED),
        ("FONTNAME",      (6,1),(6,-1),  "Helvetica-Bold"),
        ("ALIGN",         (6,1),(6,-1),  "RIGHT"),
        ("BACKGROUND",    (6,1),(6,-1),  colors.HexColor("#fff5f5")),
        ("LINEAFTER",     (5,0),(5,-1),  0.8, colors.HexColor("#aaaaaa")),
        ("ALIGN",         (4,1),(4,-1),  "CENTER"),
        ("GRID",          (0,0),(-1,-1), 0.3, GBRD),
        ("TOPPADDING",    (0,0),(-1,-1), 3),
        ("BOTTOMPADDING", (0,0),(-1,-1), 3),
        ("LEFTPADDING",   (0,0),(-1,-1), 4),
        ("RIGHTPADDING",  (0,0),(-1,-1), 4),
    ]))
    return t


def section_banner(label, n_tx, subtotal, color, styles):
    s = ParagraphStyle("sb", parent=styles["Normal"], fontSize=10,
                       fontName="Helvetica-Bold", textColor=colors.white)
    sm = ParagraphStyle("sm2", parent=styles["Normal"], fontSize=8, textColor=colors.HexColor("#cccccc"))
    tbl = Table([[
        Paragraph(label, s),
        Paragraph(f"{n_tx} transactions  |  {subtotal}", sm),
    ]], colWidths=["35%","65%"])
    tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,-1), color),
        ("TOPPADDING",    (0,0),(-1,-1), 7),
        ("BOTTOMPADDING", (0,0),(-1,-1), 7),
        ("LEFTPADDING",   (0,0),(0,-1),  12),
        ("RIGHTPADDING",  (-1,0),(-1,-1),12),
        ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
        ("ALIGN",         (1,0),(1,-1),  "RIGHT"),
    ]))
    return tbl


def subtotal_bar(label, amount, color, styles):
    s = ParagraphStyle("stb", parent=styles["Normal"], fontSize=8.5,
                       fontName="Helvetica-Bold", textColor=color)
    tbl = Table([[Paragraph(label, s), Paragraph(_fmt(amount), s)]],
                colWidths=[5.5*inch, 1.5*inch])
    tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,-1), colors.HexColor("#f0f0f0")),
        ("ALIGN",         (1,0),(1,-1),  "RIGHT"),
        ("TOPPADDING",    (0,0),(-1,-1), 4),
        ("BOTTOMPADDING", (0,0),(-1,-1), 4),
        ("LEFTPADDING",   (0,0),(0,-1),  8),
        ("RIGHTPADDING",  (-1,0),(-1,-1),8),
        ("LINEABOVE",     (0,0),(-1,0),  0.5, color),
    ]))
    return tbl


def net_bar(net, styles):
    is_pos = net >= 0
    color  = GREEN if is_pos else RED
    label  = "NET CASH FLOW (Income − Expenses):"
    s = ParagraphStyle("net", parent=styles["Normal"], fontSize=10,
                       fontName="Helvetica-Bold", textColor=colors.white)
    tbl = Table([[Paragraph(label, s), Paragraph(_fmt(abs(net)) + (" ▲" if is_pos else " ▼"), s)]],
                colWidths=[5.0*inch, 2.0*inch])
    tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,-1), color),
        ("ALIGN",         (1,0),(1,-1),  "RIGHT"),
        ("TOPPADDING",    (0,0),(-1,-1), 7),
        ("BOTTOMPADDING", (0,0),(-1,-1), 7),
        ("LEFTPADDING",   (0,0),(0,-1),  10),
        ("RIGHTPADDING",  (-1,0),(-1,-1),10),
    ]))
    return tbl


# ─── GENERADOR PDF ──────────────────────────

def generar_pdf(rows_by_month, selected_months, year):
    all_rows    = [r for rows in rows_by_month.values() for r in rows]
    income_rows = [r for r in all_rows if is_income(r)]
    expense_rows= [r for r in all_rows if is_expense(r)]
    total_in    = sum(float(r["usd"] or 0) for r in income_rows)
    total_exp   = sum(float(r["usd"] or 0) for r in expense_rows)
    net         = total_in - total_exp

    if len(selected_months) == 1:
        m = selected_months[0]
        periodo = f"{MONTHS[m]} {year}"
    else:
        names = [MONTHS[m] for m in selected_months]
        periodo = f"{names[0]} – {names[-1]} {year}"

    styles = getSampleStyleSheet()
    def S(name, parent="Normal", **kw):
        return ParagraphStyle(name, parent=styles[parent], **kw)

    t_s  = S("t",  fontSize=15, fontName="Helvetica-Bold", textColor=colors.white)
    sb_s = S("sb", fontSize=9,  textColor=colors.HexColor("#c0c0cc"))
    h_s  = S("h",  fontSize=8.5,fontName="Helvetica-Bold", textColor=MID, spaceBefore=12, spaceAfter=5)
    b_s  = S("b",  fontSize=10, textColor=colors.HexColor("#222222"), leading=16)
    sm_s = S("sm", fontSize=7.5,textColor=MID)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter,
        leftMargin=0.65*inch, rightMargin=0.65*inch,
        topMargin=0.6*inch,   bottomMargin=0.6*inch)
    story = []

    def banner(title, sub, w1="60%", w2="40%"):
        tbl = Table([[Paragraph(title, t_s), Paragraph(sub, sb_s)]], colWidths=[w1, w2])
        tbl.setStyle(TableStyle([
            ("BACKGROUND",   (0,0),(-1,-1), DARK),
            ("TOPPADDING",   (0,0),(-1,-1), 10),
            ("BOTTOMPADDING",(0,0),(-1,-1), 10),
            ("LEFTPADDING",  (0,0),(0,-1),  12),
            ("RIGHTPADDING", (-1,0),(-1,-1),12),
            ("VALIGN",       (0,0),(-1,-1), "MIDDLE"),
            ("ALIGN",        (1,0),(1,-1),  "RIGHT"),
        ]))
        return tbl

    # ── PAGE 1: LEDGER ──────────────────────
    story.append(banner("LCC — Cash Ledger (Cuba)", f"Period: {periodo}"))
    story.append(Spacer(1, 8))

    # Summary strip
    strip = Table([
        [Paragraph("Total Income (USD)", sm_s),
         Paragraph("Total Expenses (USD)", sm_s),
         Paragraph("Net Cash Flow", sm_s)],
        [Paragraph(f'<font color="#0f6e56"><b>{_fmt(total_in)}</b></font>', styles["Normal"]),
         Paragraph(f'<font color="#a32d2d"><b>{_fmt(total_exp)}</b></font>', styles["Normal"]),
         Paragraph(f'<font color="{"#0f6e56" if net>=0 else "#a32d2d"}"><b>{_fmt(abs(net))} {"▲" if net>=0 else "▼"}</b></font>', styles["Normal"])],
    ], colWidths=["33%","33%","34%"])
    strip.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,-1), GBGD),
        ("GRID",          (0,0),(-1,-1), 0.3, GBRD),
        ("TOPPADDING",    (0,0),(-1,-1), 6),
        ("BOTTOMPADDING", (0,0),(-1,-1), 6),
        ("LEFTPADDING",   (0,0),(-1,-1), 10),
        ("FONTSIZE",      (0,1),(-1,-1), 11),
    ]))
    story.append(strip)

    # ── INCOME SECTION ──
    if income_rows:
        story.append(Spacer(1, 12))
        story.append(section_banner("INCOME", len(income_rows), _fmt(total_in),
                                    colors.HexColor("#0f6e56"), styles))
        story.append(Spacer(1, 4))
        story.append(make_tx_table(income_rows, colors.HexColor("#0f6e56")))
        story.append(subtotal_bar("Total Income:", total_in, GREEN, styles))

    # ── EXPENSES SECTION ──
    if expense_rows:
        story.append(Spacer(1, 12))
        story.append(section_banner("EXPENSES", len(expense_rows), _fmt(total_exp),
                                    RED, styles))
        story.append(Spacer(1, 4))
        story.append(make_tx_table(expense_rows, RED))
        story.append(subtotal_bar("Total Expenses:", total_exp, RED, styles))

    # ── NET ──
    story.append(Spacer(1, 10))
    story.append(net_bar(net, styles))

    # ── PAGE 2: APPROVAL ────────────────────
    story.append(PageBreak())
    story.append(banner("LCC — Cash Ledger Approval", f"Period: {periodo}", "65%", "35%"))
    story.append(Spacer(1, 20))
    story.append(Paragraph("SUMMARY", h_s))

    recap = Table([
        ["Period",               periodo],
        ["Total Income (USD)",   _fmt(total_in)],
        ["Total Expenses (USD)", _fmt(total_exp)],
        ["Net Cash Flow (USD)",  _fmt(net)],
        ["No. of Transactions",  str(len(all_rows))],
    ], colWidths=[3*inch, 2.5*inch])
    recap.setStyle(TableStyle([
        ("FONTNAME",      (0,0),(-1,-1), "Helvetica"),
        ("FONTNAME",      (0,0),(0,-1),  "Helvetica-Bold"),
        ("FONTSIZE",      (0,0),(-1,-1), 10),
        ("ALIGN",         (1,0),(1,-1),  "RIGHT"),
        ("TOPPADDING",    (0,0),(-1,-1), 5),
        ("BOTTOMPADDING", (0,0),(-1,-1), 5),
        ("TEXTCOLOR",     (1,1),(1,1),   GREEN),
        ("TEXTCOLOR",     (1,2),(1,2),   RED),
        ("TEXTCOLOR",     (1,3),(1,3),   GREEN if net >= 0 else RED),
        ("FONTNAME",      (0,3),(-1,3),  "Helvetica-Bold"),
        ("LINEABOVE",     (0,3),(-1,3),  1, DARK),
        ("LINEBELOW",     (0,-1),(-1,-1),1, DARK),
    ]))
    story.append(recap)
    story.append(Spacer(1, 28))

    story.append(Paragraph("APPROVAL", h_s))
    story.append(Paragraph(
        f"Reviewed and approved. I confirm the transactions recorded in this Cash Ledger "
        f"for <b>{periodo}</b> are accurate and complete.",
        b_s))
    story.append(Spacer(1, 36))

    sig = Table([
        [Paragraph("<b>Approved by:</b>", b_s), ""],
        [Spacer(1,6), ""],
        [Paragraph("Name:", b_s), ""],
        [Spacer(1,14), ""],
        [Paragraph("Signature:", b_s), ""],
        [Spacer(1,22), ""],
        [Paragraph("Date:", b_s), ""],
    ], colWidths=[3.5*inch, 3*inch])
    sig.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,-1), GBGD),
        ("BOX",           (0,0),(-1,-1), 0.5, GBRD),
        ("LEFTPADDING",   (0,0),(-1,-1), 16),
        ("RIGHTPADDING",  (0,0),(-1,-1), 16),
        ("TOPPADDING",    (0,0),(1,0),   12),
        ("BOTTOMPADDING", (0,-1),(1,-1), 14),
        ("TOPPADDING",    (0,1),(-1,-1), 2),
        ("BOTTOMPADDING", (0,1),(-1,-1), 2),
        ("LINEBELOW",     (0,2),(0,2),   1, DARK),
        ("LINEBELOW",     (0,4),(0,4),   1, DARK),
        ("LINEBELOW",     (0,6),(0,6),   1, DARK),
    ]))
    story.append(sig)
    story.append(Spacer(1, 40))
    story.append(HRFlowable(width="100%", thickness=0.5, color=GBRD, spaceAfter=6))
    story.append(Paragraph(f"Generated: {datetime.now().strftime('%B %d, %Y')}", sm_s))

    doc.build(story)
    buf.seek(0)
    return buf


# ─── GENERADOR EXCEL ────────────────────────

def generar_excel(rows_by_month, selected_months, year):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    all_rows     = [r for rows in rows_by_month.values() for r in rows]
    income_rows  = [r for r in all_rows if is_income(r)]
    expense_rows = [r for r in all_rows if is_expense(r)]
    total_in     = sum(float(r["usd"] or 0) for r in income_rows)
    total_exp    = sum(float(r["usd"] or 0) for r in expense_rows)
    net          = total_in - total_exp

    if len(selected_months) == 1:
        m = selected_months[0]
        periodo = f"{MONTHS[m]} {year}"
    else:
        names = [MONTHS[m] for m in selected_months]
        periodo = f"{names[0]} – {names[-1]} {year}"

    wb = Workbook()

    dark_fill  = PatternFill("solid", fgColor="1A1A2E")
    green_fill = PatternFill("solid", fgColor="0F6E56")
    red_fill   = PatternFill("solid", fgColor="A32D2D")
    net_fill   = PatternFill("solid", fgColor="0F6E56" if net >= 0 else "A32D2D")
    gray_fill  = PatternFill("solid", fgColor="F5F5F5")
    gray2_fill = PatternFill("solid", fgColor="EBEBEB")
    usd_fill   = PatternFill("solid", fgColor="FFF0F0")
    inc_fill   = PatternFill("solid", fgColor="F0FFF8")
    white_fill = PatternFill("solid", fgColor="FFFFFF")

    white_bold = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    normal     = Font(name="Arial", size=9)
    mid_font   = Font(name="Arial", size=9, color="888888")
    usd_font   = Font(name="Arial", bold=True, size=9, color="A32D2D")
    inc_font   = Font(name="Arial", bold=True, size=9, color="0F6E56")
    tot_font   = Font(name="Arial", bold=True, size=10, color="FFFFFF")

    thin  = Side(style="thin",   color="DDDDDD")
    med   = Side(style="medium", color="AAAAAA")
    bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)
    usd_bdr = Border(left=med, right=thin, top=thin, bottom=thin)

    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center")
    right  = Alignment(horizontal="right",  vertical="center")

    num_usd   = "$#,##0.00"
    num_local = "###,##0.00"
    col_widths = [12, 14, 22, 48, 10, 16, 13]
    heads = ["Date", "Property", "Category", "Description", "Currency", "Local Amount", "USD"]

    def write_section(ws, section_rows, start_row, section_label, hdr_fill, usd_color, usd_fg):
        # Section header
        ws.merge_cells(f"A{start_row}:G{start_row}")
        c = ws.cell(row=start_row, column=1, value=section_label)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill = hdr_fill; c.alignment = center
        ws.row_dimensions[start_row].height = 18
        start_row += 1

        # Column headers
        for i, h in enumerate(heads, 1):
            c = ws.cell(row=start_row, column=i, value=h)
            c.font = white_bold; c.fill = dark_fill
            c.alignment = center; c.border = bdr
        ws.row_dimensions[start_row].height = 16
        start_row += 1

        usd_fill_s = PatternFill("solid", fgColor=usd_fg)
        usd_font_s = Font(name="Arial", bold=True, size=9, color=usd_color)

        for idx, r in enumerate(section_rows):
            row_num = start_row + idx
            is_alt  = idx % 2 == 1
            bg = gray_fill if is_alt else white_fill

            def cell(col, val, font=normal, fill=None, align=left, fmt=None, border=bdr):
                c = ws.cell(row=row_num, column=col, value=val)
                c.font = font; c.fill = fill or bg
                c.alignment = align; c.border = border
                if fmt: c.number_format = fmt
                return c

            try: cell(1, f"{int(r['month']):02d}/{int(r['day']):02d}/{r['year']}", align=center)
            except: cell(1, "")
            cell(2, _prop(r["prop"]))
            cell(3, r["cat"])
            cell(4, r["desc"])
            cell(5, r["cur"], align=center)
            try:
                cell(6, float(r["local"] or 0), font=mid_font, align=right, fmt=num_local)
            except: cell(6, "")
            try:
                c = ws.cell(row=row_num, column=7, value=float(r["usd"] or 0))
                c.font = usd_font_s; c.fill = usd_fill_s
                c.alignment = right; c.number_format = num_usd; c.border = usd_bdr
            except: ws.cell(row=row_num, column=7, value="")
            ws.row_dimensions[row_num].height = 15

        return start_row + len(section_rows)

    # ── Main sheet ───────────────────────────
    ws = wb.active
    ws.title = "Cash Ledger"

    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells("A1:G1")
    ws["A1"] = f"LCC — Cash Ledger (Cuba)   |   Period: {periodo}"
    ws["A1"].font = white_bold; ws["A1"].fill = dark_fill; ws["A1"].alignment = center
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:G2")
    ws["A2"] = "Expense & Income review — for Rolando's approval"
    ws["A2"].font = Font(name="Arial", size=9, color="888888", italic=True)
    ws["A2"].alignment = center; ws.row_dimensions[2].height = 16

    current_row = 4
    current_row = write_section(ws, income_rows,  current_row, "INCOME",   green_fill, "0F6E56", "F0FFF8")
    # Subtotal income
    ws.row_dimensions[current_row].height = 16
    for col in range(1, 7):
        c = ws.cell(row=current_row, column=col, value="")
        c.fill = PatternFill("solid", fgColor="E8F8F3"); c.border = bdr
    ws.cell(row=current_row, column=1, value="Total Income").font = Font(name="Arial", bold=True, size=9, color="0F6E56")
    ws.cell(row=current_row, column=1).fill = PatternFill("solid", fgColor="E8F8F3")
    ws.cell(row=current_row, column=1).alignment = center
    c = ws.cell(row=current_row, column=7, value=f"=SUM(G4:G{current_row-1})")
    c.font = Font(name="Arial", bold=True, size=9, color="0F6E56")
    c.fill = PatternFill("solid", fgColor="E8F8F3")
    c.alignment = right; c.number_format = num_usd
    current_row += 2

    exp_start = current_row
    current_row = write_section(ws, expense_rows, current_row, "EXPENSES", red_fill,   "A32D2D", "FFF0F0")
    # Subtotal expenses
    ws.row_dimensions[current_row].height = 16
    for col in range(1, 7):
        c = ws.cell(row=current_row, column=col, value="")
        c.fill = PatternFill("solid", fgColor="FFE8E8"); c.border = bdr
    ws.cell(row=current_row, column=1, value="Total Expenses").font = Font(name="Arial", bold=True, size=9, color="A32D2D")
    ws.cell(row=current_row, column=1).fill = PatternFill("solid", fgColor="FFE8E8")
    ws.cell(row=current_row, column=1).alignment = center
    c = ws.cell(row=current_row, column=7, value=f"=SUM(G{exp_start+2}:G{current_row-1})")
    c.font = Font(name="Arial", bold=True, size=9, color="A32D2D")
    c.fill = PatternFill("solid", fgColor="FFE8E8")
    c.alignment = right; c.number_format = num_usd
    current_row += 2

    # Net row
    net_color = "0F6E56" if net >= 0 else "A32D2D"
    ws.row_dimensions[current_row].height = 20
    ws.merge_cells(f"A{current_row}:F{current_row}")
    c = ws.cell(row=current_row, column=1, value="NET CASH FLOW (Income − Expenses)")
    c.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="1A1A2E"); c.alignment = center
    c = ws.cell(row=current_row, column=7, value=net)
    c.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=net_color)
    c.alignment = right; c.number_format = num_usd

    ws.freeze_panes = "A4"

    # ── Approval sheet ───────────────────────
    wa = wb.create_sheet("Approval")
    wa.column_dimensions["A"].width = 28
    wa.column_dimensions["B"].width = 22

    wa.merge_cells("A1:B1")
    c = wa.cell(row=1, column=1, value=f"LCC — Cash Ledger Approval   |   {periodo}")
    c.font = white_bold; c.fill = dark_fill; c.alignment = center
    wa.row_dimensions[1].height = 28

    for row, label, val, color in [
        (3, "Period",               periodo,         "1A1A2E"),
        (4, "Total Income (USD)",   total_in,        "0F6E56"),
        (5, "Total Expenses (USD)", total_exp,       "A32D2D"),
        (6, "Net Cash Flow (USD)",  net,             "0F6E56" if net>=0 else "A32D2D"),
        (7, "No. of Transactions",  len(all_rows),   "1A1A2E"),
    ]:
        wa.row_dimensions[row].height = 18
        c1 = wa.cell(row=row, column=1, value=label)
        c1.font = Font(name="Arial", bold=True, size=10, color="555555")
        c1.fill = gray_fill; c1.alignment = left
        c2 = wa.cell(row=row, column=2, value=val)
        c2.font = Font(name="Arial", bold=True, size=10, color=color)
        c2.fill = gray_fill; c2.alignment = right
        if isinstance(val, float): c2.number_format = num_usd

    wa.row_dimensions[8].height = 10
    wa.merge_cells("A9:B9")
    c = wa.cell(row=9, column=1,
        value=f"Reviewed and approved. I confirm the transactions recorded in this Cash Ledger for {periodo} are accurate and complete.")
    c.font = Font(name="Arial", size=10, italic=True, color="333333")
    c.alignment = Alignment(wrap_text=True, vertical="center")
    wa.row_dimensions[9].height = 32
    wa.row_dimensions[10].height = 10

    for row, label, val in [(11,"Approved by:",""), (12,"Name:",""), (13,"Signature:",""), (14,"Date:","")]:
        wa.row_dimensions[row].height = 22
        wa.cell(row=row, column=1, value=label).font = Font(name="Arial", bold=True, size=10)
        wa.cell(row=row, column=1).fill = gray2_fill
        c = wa.cell(row=row, column=2, value=val)
        c.font = Font(name="Arial", size=10); c.fill = gray2_fill
        if label in ("Name:", "Signature:", "Date:"):
            c.border = Border(bottom=Side(style="medium", color="1A1A2E"))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─── STREAMLIT UI ───────────────────────────

st.set_page_config(page_title="Cash Ledger LCC", page_icon="📊", layout="centered")
st.markdown("<style>.block-container { max-width: 780px; }</style>", unsafe_allow_html=True)

st.markdown("## 📊 Cash Ledger — LCC Cuba")
st.markdown("Upload the Excel file, select one or more months, and download the PDF or Excel for Rolando's review.")
st.divider()

uploaded_file = st.file_uploader("Upload Cash Ledger (.xlsx)", type=["xlsx"])

if uploaded_file:
    try:
        data   = leer_excel(uploaded_file)
        grupos = agrupar(data)

        if not grupos:
            st.error("No transactions found in the file.")
        else:
            year = list(grupos.keys())[0][0]
            month_options = {MONTHS[m]: m for (y, m) in sorted(grupos.keys())}

            selected_labels = st.multiselect(
                "Select month(s)",
                options=list(month_options.keys()),
                default=[list(month_options.keys())[0]]
            )

            if not selected_labels:
                st.info("Please select at least one month.")
            else:
                selected_months = [month_options[l] for l in selected_labels]
                rows_by_month   = {m: grupos[(year, m)] for m in selected_months if (year, m) in grupos}
                all_rows        = [r for rows in rows_by_month.values() for r in rows]
                income_rows     = [r for r in all_rows if is_income(r)]
                expense_rows    = [r for r in all_rows if is_expense(r)]
                total_in        = sum(float(r["usd"] or 0) for r in income_rows)
                total_exp       = sum(float(r["usd"] or 0) for r in expense_rows)
                net             = total_in - total_exp

                # Metrics
                col1, col2, col3 = st.columns(3)
                col1.metric("Total Income (USD)",   f"${total_in:,.2f}")
                col2.metric("Total Expenses (USD)", f"${total_exp:,.2f}")
                net_label = f"${abs(net):,.2f} {'▲' if net>=0 else '▼'}"
                col3.metric("Net Cash Flow", net_label, delta_color="normal")

                st.divider()

                # Preview
                if income_rows:
                    with st.expander(f"💚 Income — {len(income_rows)} transactions", expanded=False):
                        st.dataframe([{
                            "Date": _fecha(r), "Property": _prop(r["prop"]),
                            "Category": r["cat"], "Description": r["desc"],
                            "Currency": r["cur"], "Local Amount": fmt_local(r),
                            "USD": f'${float(r["usd"] or 0):,.2f}',
                        } for r in income_rows], use_container_width=True, hide_index=True)

                if expense_rows:
                    with st.expander(f"🔴 Expenses — {len(expense_rows)} transactions", expanded=True):
                        st.dataframe([{
                            "Date": _fecha(r), "Property": _prop(r["prop"]),
                            "Category": r["cat"], "Description": r["desc"],
                            "Currency": r["cur"], "Local Amount": fmt_local(r),
                            "USD": f'${float(r["usd"] or 0):,.2f}',
                        } for r in expense_rows], use_container_width=True, hide_index=True)

                st.divider()
                st.markdown("**Download files**")

                tz  = pytz.timezone("America/Toronto")
                ts  = datetime.now(tz).strftime("%Y%m%d_%H%M")
                if len(selected_months) == 1:
                    period_str = f"{MONTHS[selected_months[0]]}{year}"
                else:
                    period_str = f"{MONTHS[selected_months[0]]}-{MONTHS[selected_months[-1]]}{year}"
                base = f"SS_CashLedger_Cuba_{period_str}_{ts}"

                col_pdf, col_xlsx = st.columns(2)

                with col_pdf:
                    pdf_buf = generar_pdf(rows_by_month, selected_months, year)
                    st.download_button(
                        label="📄 Download PDF",
                        data=pdf_buf,
                        file_name=f"{base}.pdf",
                        mime="application/pdf",
                        use_container_width=True,
                    )

                with col_xlsx:
                    xl_buf = generar_excel(rows_by_month, selected_months, year)
                    st.download_button(
                        label="📊 Download Excel",
                        data=xl_buf,
                        file_name=f"{base}_Review.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )

    except Exception as e:
        st.error(f"Error reading the file: {e}")
        st.exception(e)

else:
    st.info("Upload the Excel file to get started.")
